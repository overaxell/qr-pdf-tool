import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import io
import zipfile
import requests
import qrcode
from PIL import Image
from openpyxl import load_workbook  # чтение гиперссылок из Excel
import numpy as np

# --- КОНФИГУРАЦИЯ СТРАНИЦЫ ---
st.set_page_config(
    page_title="Кюарыч",
    page_icon="▪️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# --- CSS СТИЛИ + АНИМАЦИИ ---
st.markdown(
    """
<link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;800&display=swap" rel="stylesheet">
<style>
    .stApp {background-color: #FFFFFF !important;}

    /* span не трогаем, чтобы не ломать иконки (keyboard_arrow_right и т.п.) */
    html, body, p, div, h1, h2, h3, h4, h5, h6, button, input, textarea, label, li, a {
        font-family: 'Manrope', sans-serif !important;
        color: #000000 !important;
    }

    /* АНИМАЦИИ */
    @keyframes fadeInSoft {
        0% {opacity: 0;}
        100% {opacity: 1;}
    }
    @keyframes pulseSoft {
        0%   {transform: scale(1);}
        50%  {transform: scale(1.02);}
        100% {transform: scale(1);}
    }

    /* typewriter для "Кюарыч" (7 символов) */
    @keyframes typing {
        0%   { width: 0; }
        20%  { width: 7.2ch; }  /* ~2 сек — набор */
        80%  { width: 7.2ch; }  /* пауза с полным словом */
        100% { width: 0; }      /* стирание */
    }
    @keyframes blink-caret {
        from, to { border-color: transparent; }
        50%      { border-color: #000000; }
    }

    .big-title {
        font-size: 96px !important;
        font-weight: 800 !important;
        line-height: 1.0 !important;
        letter-spacing: -3px !important;
        margin-bottom: 20px !important;
        display: inline-block !important;
        overflow: hidden !important;
        white-space: nowrap !important;
        border-right: 4px solid #000000;
        width: 7.2ch;
        animation:
            typing 10s steps(7, end) infinite,
            blink-caret .75s step-end infinite;
    }

    .description {
        font-size: 18px !important;
        color: #666 !important;
        line-height: 1.5 !important;
        margin-bottom: 18px !important;
        max-width: 650px;
        display: block !important;
        animation: fadeInSoft 0.6s ease-out both;
        animation-delay: 0.1s;
    }
    .section-title {
        font-size: 32px !important;
        font-weight: 600 !important;
        margin-top: 24px !important;
    }
    header, footer, #MainMenu {visibility: hidden;}
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 4rem !important;
        animation: fadeInSoft 0.4s ease-out;
    }

    /* === ЕДИНАЯ РАМКА ДЛЯ ВСЕХ INPUT/NUMBER_INPUT === */
    div[data-baseweb="input"] {
        border-radius: 14px !important;
        border: 1px solid #E0E0E0 !important;
        background-color: #FFFFFF !important;
        box-shadow: none !important;
        overflow: hidden !important;
        transition: border-color 0.18s ease-out,
                    box-shadow 0.18s ease-out,
                    transform 0.10s ease-out;
    }
    div[data-baseweb="input"]:focus-within {
        border-color: #000000 !important;
        box-shadow: 0 0 0 1px #00000010;
        transform: translateY(-1px);
    }

    /* Убираем любые внутренние рамки и тени (в том числе у number_input) */
    div[data-baseweb="input"] * {
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }

    /* Само текстовое поле */
    div[data-baseweb="input"] input {
        background-color: transparent !important;
    }

    /* Кнопки +/- у number_input */
    div[data-baseweb="input"] button {
        background-color: #F5F5F5 !important;
        color: #000000 !important;
    }
    div[data-baseweb="input"] button:hover {
        color: #FFFFFF !important;
    }

    div.stButton, div.stDownloadButton {
        width: 100% !important;
        display: block !important;
    }

    /* Обычные кнопки (например, Генерация) — чёрные с микроанимацией */
    div.stButton > button {
        width: 100% !important;
        min-width: 300px !important;
        background-color: #000000 !important;
        color: #FFFFFF !important;
        border-radius: 14px !important;
        padding: 18px 40px !important;
        font-size: 18px !important;
        font-weight: 500 !important;
        border: none !important;
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        white-space: nowrap !important;
        transition: background-color 0.18s ease-out,
                    transform 0.12s ease-out,
                    box-shadow 0.18s ease-out;
        box-shadow: 0 6px 14px rgba(0, 0, 0, 0.08);
    }
    div.stButton > button:hover {
        background-color: #333333 !important;
        transform: translateY(-1px) scale(1.01);
        box-shadow: 0 10px 22px rgba(0, 0, 0, 0.12);
    }
    div.stButton > button:active {
        transform: translateY(0) scale(0.99);
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.10);
    }

    /* Кнопка скачивания — маджента + пульс */
    div.stDownloadButton > button {
        width: 100% !important;
        min-width: 300px !important;
        background-color: #f0047f !important;
        color: #FFFFFF !important;
        border-radius: 14px !important;
        padding: 18px 40px !important;
        font-size: 18px !important;
        font-weight: 500 !important;
        border: none !important;
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        white-space: nowrap !important;
        box-shadow: 0 8px 18px rgba(240, 4, 127, 0.4);
        transition: background-color 0.18s ease-out,
                    transform 0.12s ease-out,
                    box-shadow 0.18s ease-out;
        animation: pulseSoft 1.6s ease-in-out infinite;
    }
    div.stDownloadButton > button:hover {
        background-color: #c00367 !important;
        transform: translateY(-1px) scale(1.01);
        box-shadow: 0 12px 26px rgba(240, 4, 127, 0.5);
    }
    div.stDownloadButton > button:active {
        transform: translateY(0) scale(0.99);
        box-shadow: 0 6px 14px rgba(240, 4, 127, 0.4);
    }

    div.stButton > button p, div.stDownloadButton > button p {
        color: #FFFFFF !important;
        margin: 0;
    }

    /* Табы: только чёрный underline у активного */
    .stTabs [data-baseweb="tab-list"] {
        border-bottom: 0 !important;
        box-shadow: none !important;
    }
    .stTabs [aria-selected="true"] {
        color: #000 !important;
        border-bottom: 2px solid #000 !important;
    }

    hr {border-color: #eee !important; margin: 24px 0 !important;}
</style>
""",
    unsafe_allow_html=True,
)

# --- КОНСТАНТЫ ---
MM_TO_POINT = 72 / 25.4
HEADERS = {"User-Agent": "Mozilla/5.0"}


def mm_to_pt(mm_val: float) -> float:
    return mm_val * MM_TO_POINT


# --- QR-ИЗОБРАЖЕНИЕ ---
def get_or_generate_qr_image(link: str):
    if not link or str(link).lower() == "nan":
        return None
    link = str(link).strip()
    if not link:
        return None

    try:
        download_url = link
        if not download_url.startswith("http"):
            download_url = "https://" + download_url
        resp = requests.get(download_url, headers=HEADERS, timeout=3)
        if resp.status_code == 200:
            pil_img = Image.open(io.BytesIO(resp.content))
            img_byte_arr = io.BytesIO()
            pil_img.save(img_byte_arr, format="PNG")
            return img_byte_arr.getvalue()
    except Exception:
        pass

    try:
        qr = qrcode.QRCode(box_size=10, border=0)
        qr.add_data(link)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format="PNG")
        return img_byte_arr.getvalue()
    except Exception:
        return None


# --- ВСПОМОГАТЕЛЬНЫЙ РАСТРОВЫЙ ДЕТЕКТОР БЕЛЫХ ОБЛАСТЕЙ ---
def _detect_white_rectangles_raster(
    pdf_bytes: bytes,
    white_threshold: int = 245,
    min_area_ratio: float = 0.001,
    max_area_ratio: float = 0.9,
):
    rects_pt = []

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        page = doc[0]
        page_w_pt = page.rect.width
        page_h_pt = page.rect.height

        pix = page.get_pixmap(alpha=False)
        img_w, img_h = pix.width, pix.height
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(img_h, img_w, 3)

    gray = img.mean(axis=2)

    mask = gray > white_threshold
    visited = np.zeros_like(mask, dtype=bool)

    img_area = img_w * img_h

    def flood_fill(sx, sy):
        stack = [(sx, sy)]
        visited[sy, sx] = True
        min_x = max_x = sx
        min_y = max_y = sy

        while stack:
            x, y = stack.pop():

            if x < min_x:
                min_x = x
            if x > max_x:
                max_x = x
            if y < min_y:
                min_y = y
            if y > max_y:
                max_y = y

            for nx, ny in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if 0 <= nx < img_w and 0 <= ny < img_h:
                    if mask[ny, nx] and not visited[ny, nx]:
                        visited[ny, nx] = True
                        stack.append((nx, ny))

        return min_x, min_y, max_x, max_y

    for y in range(img_h):
        for x in range(img_w):
            if mask[y, x] and not visited[y, x]:
                x1, y1, x2, y2 = flood_fill(x, y)

                w = x2 - x1 + 1
                h = y2 - y1 + 1
                area = w * h
                area_ratio = area / img_area
                if area_ratio < min_area_ratio or area_ratio > max_area_ratio:
                    continue

                aspect = w / h if h != 0 else 0
                if aspect < 0.5 or aspect > 2.0:
                    continue

                x_pt = x1 * page_w_pt / img_w
                y_pt = y1 * page_h_pt / img_h
                w_pt = w * page_w_pt / img_w
                h_pt = h * page_h_pt / img_h

                rects_pt.append((x_pt, y_pt, w_pt, h_pt))

    rects_pt.sort(key=lambda r: r[2] * r[3], reverse=True)
    return rects_pt


# --- ДЕТЕКТОР БЕЛЫХ КВАДРАТОВ В PDF ---
def detect_white_rectangles_in_pdf(pdf_bytes: bytes):
    rects_pt = []

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        page = doc[0]
        page_rect = page.rect
        page_w_pt = page_rect.width
        page_h_pt = page_rect.height

        drawings = page.get_drawings()

        for d in drawings:
            r = d.get("rect", None)
            if r is None:
                continue

            fill = d.get("fill", None)
            if fill is None:
                fill = d.get("color", None)
            if fill is None:
                continue

            fr, fg, fb = fill
            if fr < 0.95 or fg < 0.95 or fb < 0.95:
                continue

            w_pt = r.width
            h_pt = r.height
            if w_pt <= 0 or h_pt <= 0:
                continue

            area = w_pt * h_pt
            area_ratio = area / (page_w_pt * page_h_pt)
            if area_ratio < 0.001 or area_ratio > 0.6:
                continue

            aspect = w_pt / h_pt if h_pt != 0 else 0
            if aspect < 0.5 or aspect > 2.0:
                continue

            rects_pt.append((r.x0, r.y0, w_pt, h_pt))

    if rects_pt:
        rects_pt.sort(key=lambda r: r[2] * r[3], reverse=True)
        return rects_pt

    return _detect_white_rectangles_raster(pdf_bytes)


# --- ОБРАБОТКА PDF И ГЕНЕРАЦИЯ ZIP ---
def process_files(pdf_file, links, p_name, p_size, mode, x_mm, y_mm, size_mm):
    zip_buffer = io.BytesIO()
    pdf_file.seek(0)
    pdf_bytes = pdf_file.read()
    success_count = 0
    errors_log = []
    total_links = len(links)

    white_rects = []
    if mode == "white_rect":
        try:
            white_rects_raw = detect_white_rectangles_in_pdf(pdf_bytes)
            min_size_mm = 25.0
            white_rects = [
                r for r in white_rects_raw if min(r[2], r[3]) / MM_TO_POINT >= min_size_mm
            ]
            if not white_rects:
                errors_log.append("Белый квадрат не найден или его сторона меньше 25 мм.")
                return None, errors_log
        except Exception as e:
            errors_log.append(f"Автообнаружение: ошибка {e}")
            return None, errors_log

    my_bar = st.progress(0, text="Начинаем обработку...")

    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for i, url in enumerate(links, start=1):
            my_bar.progress(i / total_links, text=f"Обработка {i} из {total_links}")
            try:
                filename = f"{p_name}_{p_size}_{i:02d}.pdf"
                qr_bytes = get_or_generate_qr_image(url)

                if qr_bytes:
                    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                        page = doc[0]

                        if mode == "white_rect" and white_rects:
                            rx, ry, rw, rh = white_rects[0]
                            margin_pt = mm_to_pt(2.0)
                            inner_w = rw - 2 * margin_pt
                            inner_h = rh - 2 * margin_pt
                            qr_size_pt = min(inner_w, inner_h)

                            if qr_size_pt <= 0:
                                errors_log.append(
                                    "Подходящий квадрат найден, но внутренняя область слишком маленькая."
                                )
                                return None, errors_log

                            x_pt = rx + margin_pt + (inner_w - qr_size_pt) / 2
                            y_pt = ry + margin_pt + (inner_h - qr_size_pt) / 2
                        else:
                            x_pt = mm_to_pt(x_mm)
                            y_pt = mm_to_pt(y_mm)
                            qr_size_pt = mm_to_pt(size_mm)

                        rect = fitz.Rect(
                            x_pt,
                            y_pt,
                            x_pt + qr_size_pt,
                            y_pt + qr_size_pt,
                        )
                        page.insert_image(rect, stream=qr_bytes)

                        pdf_out = doc.tobytes()
                        zf.writestr(filename, pdf_out)
                        success_count += 1
                else:
                    errors_log.append(
                        f"Ссылка №{i}: Пустые данные или сбой при создании QR"
                    )
            except Exception as e:
                errors_log.append(f"Ссылка №{i}: Ошибка {e}")

    my_bar.empty()
    zip_buffer.seek(0)

    if success_count == 0:
        return None, errors_log
    return zip_buffer, errors_log


# --- ЧТЕНИЕ ССЫЛОК ИЗ EXCEL ---
def extract_links_from_excel(file) -> list:
    file_bytes = file.read()
    bio = io.BytesIO(file_bytes)

    wb = load_workbook(bio, data_only=True)
    ws = wb.active

    all_columns = []
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(1, max_col + 1):
        col_values = []
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell.hyperlink and cell.hyperlink.target:
                text = str(cell.hyperlink.target).strip()
            else:
                val = cell.value
                text = "" if val is None else str(val).strip()

            col_values.append(text)
        all_columns.append(col_values)

    import re

    def is_url_like(s: str) -> bool:
        s = s.strip()
        if len(s) <= 5:
            return False
        return bool(re.search(r"http|www|\.[a-zA-Z]{2,}", s))

    best_idx = None
    best_score = 0

    for idx, col_vals in enumerate(all_columns):
        if not col_vals:
            continue
        score = sum(1 for v in col_vals if is_url_like(v))
        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx is None or best_score == 0:
        return []

    col_vals = all_columns[best_idx]

    if col_vals and not is_url_like(col_vals[0]):
        data_vals = col_vals[1:]
    else:
        data_vals = col_vals

    clean_links = [
        v.strip()
        for v in data_vals
        if v and v.strip() and v.strip().lower() not in ("nan", "none")
    ]

    return clean_links


# --- ВЕРСТКА ---
col_left, col_spacer, col_right = st.columns([1.2, 0.1, 1.1])

# ЛЕВАЯ КОЛОНКА
with col_left:
    st.markdown(
        """
    <div>
        <div class="big-title">Кюарыч</div>
        <div class="description">
            Удобный помощник для маркетинг-команды. Загружайте макет,
            вставляйте ссылки — а я красиво и точно расставлю QR-коды сам.
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown("<hr>", unsafe_allow_html=True)

    st.markdown(
        '<div class="section-title">Как назвать файл?</div>',
        unsafe_allow_html=True,
    )
    c1, c2 = st.columns(2)
    with c1:
        partner_name = st.text_input("Имя партнера", placeholder="Partner")
    with c2:
        size_name = st.text_input("Размер файла", placeholder="0x0")

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown(
        '<div class="section-title">Куда вставить QR?</div>',
        unsafe_allow_html=True,
    )

    mode = st.radio(
        "Режим позиционирования QR",
        ["Автообнаружение", "Ручная настройка координат"],
        index=0,
    )

    if mode == "Автообнаружение":
        pos_mode = "white_rect"
        st.markdown(
            """
<div style="border-radius: 12px; padding: 14px 16px; background-color:#F0F2FF; font-size:14px;">
QR будет вставлен в найденный на макете белый квадрат с отступом 2 мм.<br>
Если подходящий квадрат меньше 25 мм, коды не будут вставлены.
</div>
""",
            unsafe_allow_html=True,
        )
        x_mm = y_mm = size_mm = 0.0
    else:
        pos_mode = "manual"
        g1, g2, g3 = st.columns(3)
        with g1:
            x_mm = st.number_input("Отступ слева (мм)", value=20.0)
        with g2:
            y_mm = st.number_input("Отступ сверху (мм)", value=20.0)
        with g3:
            size_mm = st.number_input("Размер QR (мм)", value=20.0)

# ПРАВАЯ КОЛОНКА
with col_right:
    st.write("")
    st.write("")
    st.markdown(
        '<div class="description" style="margin-bottom:10px;">Источник ссылок QR</div>',
        unsafe_allow_html=True,
    )

    if "links_final" not in st.session_state:
        st.session_state.links_final = []

    tab_manual, tab_excel = st.tabs(["Вручную", "Из excel"])

    with tab_manual:
        st.write("")
        manual_text = st.text_area(
            "Ссылки списком",
            height=150,
            label_visibility="collapsed",
            placeholder="https://",
        )
        if manual_text:
            st.session_state.links_final = [
                l.strip() for l in manual_text.split("\n") if l.strip()
            ]

    with tab_excel:
        st.write("")
        uploaded_excel = st.file_uploader(
            "Excel", type=["xlsx"], key="xls", label_visibility="collapsed"
        )
        if uploaded_excel:
            try:
                uploaded_excel.seek(0)
                links_from_excel = extract_links_from_excel(uploaded_excel)
                st.session_state.links_final = links_from_excel

                if len(links_from_excel) > 0:
                    st.success(f"✅ Найдено ссылок: {len(links_from_excel)}")
                    st.markdown(
                        "<div style='height:8px;'></div>",
                        unsafe_allow_html=True,
                    )
                    with st.expander("Показать найденные ссылки", expanded=False):
                        for i, link in enumerate(links_from_excel, start=1):
                            st.write(f"{i}. {link}")
                else:
                    st.warning(
                        "Не удалось найти ссылки в файле. Проверьте, что в колонке есть URL или гиперссылки."
                    )
            except Exception as e:
                st.error(f"Ошибка файла: {e}")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div class="description" style="margin-bottom:10px;">Источник макета</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="section-title" style="margin-top:0;">Загрузите дизайн</div>',
        unsafe_allow_html=True,
    )

    uploaded_pdf = st.file_uploader(
        "PDF", type=["pdf"], key="pdf", label_visibility="collapsed"
    )

    if "prev_pdf_name" not in st.session_state:
        st.session_state.prev_pdf_name = None

    current_pdf_name = uploaded_pdf.name if uploaded_pdf is not None else None
    if current_pdf_name != st.session_state.prev_pdf_name:
        st.session_state.prev_pdf_name = current_pdf_name
        st.session_state.zip_result = None
        st.session_state.zip_name = None

    st.markdown("<br>", unsafe_allow_html=True)

    if "zip_result" not in st.session_state:
        st.session_state.zip_result = None
        st.session_state.zip_name = None

    if st.session_state.zip_result is None:
        if st.button("Генерация"):
            if not uploaded_pdf:
                st.toast("Нужен PDF!", icon="⚠️")
            elif not st.session_state.links_final:
                st.toast("Нужны ссылки!", icon="⚠️")
            elif not partner_name.strip() or not size_name.strip():
                st.toast(
                    "Сначала задайте имя партнера и размер файла.",
                    icon="⚠️",
                )
            else:
                p_n = partner_name.strip()
                s_n = size_name.strip()

                res, errs = process_files(
                    uploaded_pdf,
                    st.session_state.links_final,
                    p_n,
                    s_n,
                    pos_mode,
                    x_mm,
                    y_mm,
                    size_mm,
                )

                if res:
                    st.session_state.zip_result = res
                    st.session_state.zip_name = f"{p_n}_{s_n}.zip"
                    st.rerun()
                else:
                    if errs:
                        st.toast(errs[0], icon="⚠️")
                    st.error("Ошибка. Проверьте ссылки или макет.")
                    if errs:
                        for e in errs:
                            st.write(e)
    else:
        st.download_button(
            "Скачать архив",
            st.session_state.zip_result,
            st.session_state.zip_name or "qrs.zip",
            "application/zip",
        )
        st.caption(
            "После нажатия дождитесь начала загрузки и не нажимайте\n"
            "кнопку несколько раз подряд."
        )
